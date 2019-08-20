from django.db.models import Q
from django.shortcuts import render, redirect, reverse
from registration.models import *
from django.core.paginator import Paginator
import openpyxl


# 主页
def index(request):
    return render(request, 'doctor/index.html')


# 信息查询
def find_0(request):
    pname = request.POST.get("pname")
    request.session["id"] = pname
    return redirect("query")


def query(request):
    ids = request.session.get("id")
    doctor = request.session.get("doctor_id")
    if ids:

        if PatientOne.objects.filter(id=ids, fk_patient_doctor_id=doctor).exists():
            messages = PatientOne.objects.filter(id=ids, patient_over_look_doctor=0, fk_patient_doctor_id=doctor)

            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages": result})
        elif PatientOne.objects.filter(patient_phone=ids, fk_patient_doctor_id=doctor).exists():
            messages = PatientOne.objects.filter(patient_phone=ids, fk_patient_doctor_id=doctor,
                                                 patient_over_look_doctor=0)

            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages": result})
        else:
            messages = PatientOne.objects.filter(patient_card=ids, fk_patient_doctor_id=doctor,
                                                 patient_over_look_doctor=0)

            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages": result})


    else:
        messages = PatientOne.objects.filter(patient_over_look_doctor=0, fk_patient_doctor_id=doctor)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages": result})


# 指定病人病症的信息的查询并分页
def patient_list_1(request, ids):
    if ids:
        messages_1 = PatientOne.objects.values("patient_status").distinct()
        messages = PatientOne.objects.filter(patient_status=ids, patient_over_look_doctor=0)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages_1": messages_1, "messages": result})
    else:
        messages_1 = PatientOne.objects.values("patient_status").distinct()
        messages = PatientOne.objects.filter(patient_over_look_doctor=0)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages_1": messages_1, "messages": result})


# 所有病人信息的查询并分页
def patient_list(request):
    messages_1 = PatientOne.objects.values("patient_status").distinct()
    doctor = request.session.get('doctor_id')
    messages = PatientOne.objects.filter(patient_over_look_doctor=0, fk_patient_doctor_id=doctor)
    limit = 3
    paginator = Paginator(messages, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    return render(request, 'doctor/patient_list.html', {"messages_1": messages_1, "messages": result})


# 病症详情
def detail(request, ids):
    infos = PatientOne.objects.get(id=ids)
    return render(request, "doctor/detail.html", {"详情": infos})


# 就诊
def dispensing(request, ids):
    us_er = PatientOne.objects.get(id=ids)
    # a = us_er.patient_is_live
    yao_pin = Medicine.objects.all()  # 后期需要修改
    medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
    paginator = Paginator(yao_pin, 5)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    paginator_1 = Paginator(medicine_one, 5)  # 按每页10条分页
    page_1 = request.GET.get('page_1', '1')  # 默认跳转到第一页
    result_1 = paginator_1.page(page_1)
    return render(request, "doctor/dispensing.html", {"us_er": us_er, "yao_pin": result, "yao_fang": result_1})


# 开药方
def Medicine_one_add(request, ids):
    if request.method == "GET":
        us_er = PatientOne.objects.get(id=ids)
        yao_pin = Medicine.objects.all()  # 后期需要修改
        # medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
        limit = 5
        paginator = Paginator(yao_pin, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, "doctor/dispensing.html", {"us_er": us_er, "yao_pin": result})
    else:
        medicine_name = request.POST.get("ypmz")
        if Medicine.objects.filter(medicine_name=medicine_name).exists():
            patient_one1 = Medicine.objects.get(medicine_name=medicine_name)
            patient_one = patient_one1.medicine_outer_price
            medicine_numbers = request.POST.get("ypsl")
            if medicine_numbers and int(medicine_numbers) < patient_one1.medicine_number:
                patient_one1.medicine_number -= int(medicine_numbers)
                patient_one1.save()
                Medicine_one.objects.create(medicine_one_name=medicine_name,
                                            medicine_one_number=medicine_numbers,
                                            medicine_one_outer_price=patient_one,
                                            fk_patient_medicine_one_id=ids, )
                return redirect(reverse(dispensing, kwargs={"ids": ids}))
            else:
                return redirect(reverse(dispensing, kwargs={"ids": ids}))
        else:
            return redirect(reverse(dispensing, kwargs={"ids": ids}))


# 撤回药方中的药品
def delete_yaofan(request, ids):
    user = Medicine_one.objects.get(id=ids)
    if ids:
        delete_1 = Medicine_one.objects.get(id=ids)
        patient_one1 = Medicine.objects.get(medicine_name=delete_1.medicine_one_name)
        patient_one1.medicine_number += delete_1.medicine_one_number
        patient_one1.save()
        delete_1.delete()
        ids = user.fk_patient_medicine_one_id
        return redirect(reverse(dispensing, kwargs={"ids": ids}))
    else:
        return redirect(reverse(dispensing, kwargs={"ids": ids}))


# # 住院
def hospital(request, ids):
    if request.method == "GET":
        us_er = PatientOne.objects.get(id=ids)
        yao_pin = Medicine.objects.all()  # 后期需要修改
        # medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
        limit = 5
        paginator = Paginator(yao_pin, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, "doctor/dispensing.html", {"us_er": us_er, "yao_pin": result})
    else:
        user = PatientOne.objects.get(id=ids)
        day = request.POST.get("day")
        if 0 <= int(day) < 10000 and day:  # 缺少条件，先住再续
            user.patient_is_live = 1
            user.patient_live_day = day
            user.save()
            return redirect(reverse(dispensing, kwargs={"ids": ids}))
        else:
            return redirect(reverse(dispensing, kwargs={"ids": ids}))


# # 药方输出到excle中
def save_excle(request, ids):
    medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
    patientone = PatientOne.objects.get(id=ids)

    one = patientone.patient_name
    a = openpyxl.Workbook()  # 打开文件
    b = a.active  # 激活表格
    b.title = "%s" % one  # 添加工作簿名称
    e = [["病人", "病症", "主治大夫", "住院天数"],
         [patientone.patient_name, patientone.patient_status, patientone.patient_main_doctor,
          patientone.patient_live_day],
         ]
    for i in range(len(e)):
        for j in range(len(e[i])):
            b.cell(i + 1, j + 1, e[i][j])  # 写入单元格
    e1 = [["药名", "价钱", "数量", ]]
    for i in medicine_one:
        f = [i.medicine_one_name, i.medicine_one_outer_price, i.medicine_one_number]  # 添加内容
        e1.append(f)
        for a in range(len(e1)):
            for j in range(len(e1[a])):
                b.cell(a + 1, j + 5, e1[a][j])  # 写入单元格
    a.save(filename='d:\\one_1.xlsx')  # 保存文件
    a.close()
    return redirect(reverse(dispensing, kwargs={"ids": ids}))


# 提交并返回主页
def submit(request, ids):
    user = PatientOne.objects.get(id=ids)
    medicine = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
    if medicine:
        medicine_one_yes_no = Medicine_one.objects.filter(fk_patient_medicine_one=ids)
        medicine_one_yes_no.update(medicine_one_yes_no=1)
        user.patient_over_look_doctor = 1
        user.save()
        doctor = request.session.get('doctor_id')
        info = Doctor.objects.get(id=doctor)
        info.waiting_number -= 1
        info.save()
        return redirect("patient_list")
    else:
        return redirect("patient_list")


# 密码
def doctor_password(request):
    return render(request, 'doctor/doctor_password.html')
