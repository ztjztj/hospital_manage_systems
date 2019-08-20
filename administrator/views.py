from django.shortcuts import render, redirect, reverse, HttpResponse
from django.shortcuts import render
from registration.models import *
from django.db.models import F, Q
import re, openpyxl
import os
from openpyxl import Workbook
from io import BytesIO
from xlwt import *
from django.http import HttpResponse


def administrator_index(request):
    '''
    管理员首页
    :param request:
    :return:
    '''
    return render(request, 'administrator/index.html')


def doctor_manage(request):
    '''
    医生管理
    :param request:
    :return:
    '''

    departments = Department.objects.all()
    users = UserInfo.objects.exclude(doctor_belong_id=None)
    # department = Department.objects.get(id=id)
    return render(request, 'administrator/doctor_manage.html',
                  {'departments': departments,
                   'users': users
                   }
                  )


def doctor_manage_type_look(request, id):
    '''
    医生管理
    :param request:
    :return:
    '''
    print("*" * 100)
    # department = Department.objects.get(id=id)
    # users=UserInfo.objects.filter(doctor_belong_id=id)
    # print('department',department)
    # return render(request, 'administrator/doctor_manage.html',
    #               {'department': department,
    #                'users': users,
    #                }
    #               )

    if id:
        users = UserInfo.objects.filter(doctor_belong_id=id)
        departments = Department.objects.all()
        # department = Department.objects.get(id=id)
        return render(request, 'administrator/doctor_manage.html',
                      {'departments': departments,
                       'users': users,
                       }
                      )
    else:
        departments = Department.objects.all()
        return render(request, 'administrator/doctor_manage.html',
                      {'departments': departments}
                      )

    # ★★★★待完成
    # def doctor_manage_excel(request):
    #     '''
    #     医生管理:所有医生信息-->导出Excel
    #     :param request:
    #     :return:
    #     '''
    #     departments = Department.objects.all()
    #     for department in departments:
    #         print(department.depart_name)
    #         a = UserInfo.objects.filter(doctor_belong__depart_name=department.depart_name)


def doctor_manage_add(request):
    '''
    医生管理：添加
    :param request:
    :return:
    '''
    return render(request, 'administrator/doctor_manage_add.html', )


def doctor_manage_del(request, user_id):
    '''
    用户管理：删除用户信息
    :param request:
    :param user_id:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.delete()
    return redirect('doctor_manage')


def doctor_manage_del_box(request):
    '''
    当选中用户时，删除用户
    :param request:
    :return:
    '''
    data = {}
    ids = request.POST.get('ids')
    ids = ids.split(',')
    print(ids)
    for i in ids:
        print(i)
        user = UserInfo.objects.filter(id=int(i))
        print('user', user)
        user.delete()
    data['status'] = '删除成功'
    return JsonResponse(data)
    # return redirect('doctor_manage')


def doctor_start_use(request, user_id):
    '''
    启用医生
    :param request:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.user_status = '启用'
    user.save()
    return redirect('doctor_manage')


def doctor_stop_use(request, user_id):
    '''
    禁用医生
    :param request:
    :param user_id:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.user_status = '禁用'
    user.save()
    return redirect('doctor_manage')
    # return redirect('start_and_stop')


def doctor_manage_look_excel(request, user_id):
    '''
    医生管理:详情-->导出Excel
    :param request:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.id = str(user.id)
    role = Role.objects.get(id=user.user_role_id)
    # print(role)
    department = Department.objects.get(id=user.doctor_belong_id)
    if user:
        wb = Workbook(encoding="utf-8")
        w = wb.add_sheet(u"医生详情表")
        w.write(0, 0, u"编号")
        w.write(1, 0, u"姓名")
        w.write(2, 0, u"性别")
        w.write(3, 0, u"年龄")
        w.write(4, 0, u"学历")
        w.write(5, 0, u"角色")
        w.write(6, 0, u"所属科室")
        w.write(7, 0, u"状态")
        w.write(8, 0, u"身份证号")
        w.write(9, 0, u"电话号码")
        w.write(10, 0, u"电子邮箱")
        w.write(0, 1, user.id)
        w.write(1, 1, user.user_name)
        w.write(2, 1, user.user_gender)
        w.write(3, 1, user.user_age)
        w.write(4, 1, user.user_level)
        w.write(5, 1, role.role_name)
        w.write(6, 1, department.depart_name)
        w.write(7, 1, user.user_status)
        w.write(8, 1, user.user_card)
        w.write(9, 1, user.user_phone)
        w.write(10, 1, user.user_email)
        exits_file = os.path.exists('doctor_manage_look.xls')
        if exits_file:
            os.remove(r"doctor_manage_look.xls")
        wb.save('doctor_manage_look.xls')
        sio = BytesIO()
        wb.save(sio)
        sio.seek(0)
        response = HttpResponse(sio.getvalue(), content_type='pplication/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=doctor_manage_look.xls'
        response.write(sio.getvalue())
        return response


def doctor_manage_search(request):
    '''
    门诊医生管理：搜索
    :param request:
    :return:
    '''
    name = request.POST.get('d_content')
    # print('*'*100)
    # print('name',name)
    if UserInfo.objects.filter(user_name=name).exists():  # 姓名
        users = UserInfo.objects.filter(user_name=name)
        # print('查询_通过姓名', users)
        # print('*' * 100)
        # for user in users:
        #     print('查询_user', user)
        #     print('查询_姓名', user.user_name)
        #     print('查询_角色', user.user_role)
        #     print('查询_所属科室', user.doctor_belong)
        #     print('查询_通过姓名', user.doctor_belong_id)
        return render(request, 'administrator/doctor_manage.html',
                      {'users': users,
                       'name': name,
                       # 'departments':departments
                       })
    elif UserInfo.objects.filter(user_gender=name).exists():  # 性别
        users = UserInfo.objects.filter(user_name=name)
        # departments = Department.objects.get(id=users.doctor_belong_id)
        # print('*' * 100)
        # print('查询_通过性别', users)
        return render(request, 'administrator/doctor_manage.html',
                      {
                          'users': users,
                          'name': name,
                          # 'departments':departments
                      })
    elif UserInfo.objects.filter(doctor_belong__depart_name=name).exists():  # 所属科室
        users = UserInfo.objects.filter(doctor_belong__depart_name=name)
        # departments = Department.objects.get(id=users.doctor_belong_id)
        # print('*' * 100)
        # print('查询_通过科室', users)
        return render(request, 'administrator/doctor_manage.html',
                      {'users': users,
                       'name': name,
                       # 'departments':departments
                       })
    else:
        return redirect('doctor_manage')


def user_manage_search(request):
    '''
    门诊医生管理：搜索
    :param request:
    :return:
    '''
    name = request.POST.get('u_content')
    # print('*'*100)
    # print('name',name)
    if UserInfo.objects.filter(user_name=name).exists():  # 姓名
        users = UserInfo.objects.filter(user_name=name)
        # for user in users:
        # user_doctor_belong_id=user.doctor_belong_id
        # departments=Department.objects.get(id=user_doctor_belong_id)
        # departments = Department.objects.get(id=users.doctor_belong_id)

        print('查询_通过姓名', users)
        print('*' * 100)
        # for user in users:
        #     print('查询_user', user)
        #     print('查询_姓名', user.user_name)
        #     print('查询_角色', user.user_role)
        #     print('查询_所属科室', user.doctor_belong)
        #     print('查询_通过姓名', user.doctor_belong_id)
        return render(request, 'administrator/user_manage.html',
                      {'users': users,
                       'name': name,
                       # 'departments':departments
                       })
    elif UserInfo.objects.filter(user_gender=name).exists():  # 性别
        users = UserInfo.objects.filter(user_name=name)
        # departments = Department.objects.get(id=users.doctor_belong_id)
        # print('*' * 100)
        # print('查询_通过性别', users)
        return render(request, 'administrator/user_manage.html',
                      {
                          'users': users,
                          'name': name,
                          # 'departments':departments
                      })
    elif UserInfo.objects.filter(doctor_belong__depart_name=name).exists():  # 所属科室
        users = UserInfo.objects.filter(doctor_belong__depart_name=name)
        # departments = Department.objects.get(id=users.doctor_belong_id)
        # print('*' * 100)
        # print('查询_通过科室', users)
        return render(request, 'administrator/user_manage.html',
                      {'users': users,
                       'name': name,
                       # 'departments':departments
                       })
    else:
        return redirect('user_manage')


def doctor_manage_look(request, user_id):
    '''
    医生管理：查看
    :param request:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    # print(user)
    role = Role.objects.get(id=user.user_role_id)
    # print(role)
    department = Department.objects.get(id=user.doctor_belong_id)
    # print(department)
    return render(request, 'administrator/doctor_manage_look.html',
                  {
                      'user': user,
                      'role': role,
                      'department': department,
                  })


def doctor_manage_edit(request, user_id):
    '''
    医生管理：编辑，修改
    :param request:
    :return:
    '''
    if request.method == 'GET':
        user = UserInfo.objects.get(id=user_id)
        department = Department.objects.get(id=user.doctor_belong_id)
        return render(request, 'administrator/doctor_manage_edit.html',
                      {
                          'user': user,
                          'department': department,
                      })
    else:
        # user = UserInfo.objects.get(id=user_id)
        user_name = request.POST.get('user_name')

        user_gender = request.POST.get('user_gender')

        user_age = request.POST.get('user_age')

        user_level = request.POST.get('user_level')

        # user_role = request.POST.get('user_role')

        user_depart = request.POST.get('user_depart')

        user_status = request.POST.get('user_status')

        user_card = request.POST.get('user_card')

        user_phone = request.POST.get('user_phone')

        user_email = request.POST.get('user_email')

        now_user = UserInfo.objects.get(id=user_id)

        now_user.user_name = user_name

        now_user.user_gender = user_gender

        now_user.user_age = user_age

        now_user.user_level = user_level

        now_user.user_status = user_status

        now_user.user_card = user_card

        now_user.user_phone = user_phone

        now_user.user_email = user_email

        department = Department.objects.get(depart_name=user_depart)
        now_user.doctor_belong_id = department.id

        now_user.save()
        # role.save()
        department.save()

        # print('用户输入的姓名：', user_name)
        #
        # print('用户输入的性别：', user_gender)
        #
        # print('用户输入的年龄：', user_age)
        #
        # print('用户输入的学历：', user_level)
        #
        # print('用户输入的角色：', user_role)
        #
        # print('用户输入的所属科室：', user_depart)
        #
        # print('用户输入的状态：', user_status)
        #
        # print('用户输入的身份证号：', user_card)
        #
        # print('用户输入的手机：', user_phone)
        #
        # print('用户输入的邮箱：', user_email)
        #
        # print('用户修改后_用户表信息：', now_user)
        #
        # print('用户修改后_用户名：', now_user.user_name)
        #
        # print('用户修改后_用户名：', now_user.user_gender)
        #
        # print('用户修改后_用户名：', now_user.user_age)
        #
        # print('用户修改后_用户名：', now_user.user_level)
        #
        # print('用户修改后_用户名：', now_user.user_status)
        # print('用户修改后_用户名：', now_user.user_card)
        #
        # print('用户修改后_用户名：', now_user.user_phone)
        #
        # print('用户修改后_用户名：', now_user.user_email)
        #
        # print('用户修改后_角色id：', now_user.user_role_id)
        #
        # print('用户修改后_所属科室id：', now_user.doctor_belong_id)
    return redirect(reverse('doctor_manage_look', kwargs={'user_id': now_user.id}))


def user_manage(request):
    '''
    用户管理
    :param request:
    :return:
    '''
    roles = Role.objects.exclude(role_name='医生')
    users = UserInfo.objects.filter(doctor_belong_id=None)
    print('roles', roles)
    print('users', users)
    return render(request, 'administrator/user_manage.html',
                  {'roles': roles,
                   'users': users,
                   })


def user_manage_type_look(request, id):
    '''
    用户管理：分角色显示角色信息
    :param request:
    :return:
    '''
    if id:
        users = UserInfo.objects.filter(user_role_id=id)
        roles = Role.objects.exclude(role_name='医生')
        # department = Department.objects.get(id=id)
        return render(request, 'administrator/user_manage.html',
                      {'users': users,
                       'roles': roles,
                       }
                      )
    else:
        roles = Role.objects.all()
        return render(request, 'administrator/doctor_manage.html',
                      {'roles': roles}
                      )


def user_manage_look_excel(request, user_id):
    '''
    用户管理:详情-->导出Excel
    :param request:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.id = str(user.id)
    role = Role.objects.get(id=user.user_role_id)
    # print(role)
    # department = Department.objects.get(id=user.doctor_belong_id)
    if user:
        wb = Workbook(encoding="utf-8")
        # wb = Workbook()
        w = wb.add_sheet(u"医生详情表")
        w.write(0, 0, u"编号")
        w.write(1, 0, u"姓名")
        w.write(2, 0, u"性别")
        w.write(3, 0, u"年龄")
        w.write(4, 0, u"学历")
        w.write(5, 0, u"角色")
        w.write(6, 0, u"状态")
        w.write(7, 0, u"身份证号")
        w.write(8, 0, u"电话号码")
        w.write(9, 0, u"电子邮箱")
        w.write(0, 1, user.id)
        w.write(1, 1, user.user_name)
        w.write(2, 1, user.user_gender)
        w.write(3, 1, user.user_age)
        w.write(4, 1, user.user_level)
        w.write(5, 1, role.role_name)
        w.write(6, 1, user.user_status)
        w.write(7, 1, user.user_card)
        w.write(8, 1, user.user_phone)
        w.write(9, 1, user.user_email)
        exits_file = os.path.exists('user_manage_look.xls')
        if exits_file:
            os.remove(r"user_manage_look.xls")
        wb.save('user_manage_look.xls')
        sio = BytesIO()
        wb.save(sio)
        sio.seek(0)
        response = HttpResponse(sio.getvalue(), content_type='pplication/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=user_manage_look.xls'
        response.write(sio.getvalue())
        return response


def user_manage_look(request, user_id):
    '''
    用户管理：添加用户
    :param request:
    :return:
    '''
    # user_role = Role.objects.get(id=)
    user = UserInfo.objects.get(id=user_id)
    role = Role.objects.get(id=user.user_role_id)
    # Role.objects.get(id=user.user_role_id)
    # print('用户角色', user.user_role_id)
    # print('用户所属科室', user.doctor_belong_id)
    # print('user', user)
    # print('role',role)
    return render(request, 'administrator/user_manage_look.html',
                  {
                      'user': user,
                      'role': role,
                  })


def user_manage_add(request):
    #     '''
    #     用户管理：添加用户
    #     :param request:
    #     :return:
    #     '''
    return render(request, 'administrator/user_manage_add.html')


def user_manage_edit(request, user_id):
    '''
    用户管理：修改用户信息
    :param request:
    :param user_id:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    # print('user',user)
    if request.method == 'GET':
        role = Role.objects.get(id=user.user_role_id)
        return render(request, 'administrator/user_manage_edit.html',
                      {
                          'user': user,
                          'role': role,
                      })
    else:
        # user = UserInfo.objects.get(id=user_id)
        user_name = request.POST.get('user_name')
        print('user_name', user_name)
        user_gender = request.POST.get('user_gender')
        user_age = request.POST.get('user_age')
        user_level = request.POST.get('user_level')
        user_status = request.POST.get('user_status')
        print('user_status', user_status)
        user_card = request.POST.get('user_card')
        user_phone = request.POST.get('user_phone')
        user_email = request.POST.get('user_email')
        user_role = request.POST.get('user_role')
        now_user = UserInfo.objects.get(id=user_id)
        # print('now_user',now_user)
        now_user.user_name = user_name
        now_user.user_gender = user_gender
        now_user.user_age = user_age
        now_user.user_level = user_level
        now_user.user_status = user_status
        now_user.user_card = user_card
        now_user.user_phone = user_phone
        now_user.user_email = user_email
        role = Role.objects.get(role_name=user_role)
        now_user.user_role_id = role.id
        # print('用户表当前的状态：',now_user.user_status)
        # if now_user.user_status=='启用':
        #     now_user.user_status=user_status
        #     print('如果是启用，禁用:',now_user.user_status)
        # else:
        #     now_user.user_status =user_status
        #     print('如果是禁用，启用：', now_user.user_status)

        now_user.save()
        role.save()
        return redirect(reverse('user_manage_look', kwargs={'user_id': now_user.id}))


def user_manage_del(request, user_id):
    '''
    用户管理：删除用户信息
    :param request:
    :param user_id:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.delete()
    # user.save()
    # return redirect(reverse('user_manage_look',kwargs={'user_id':user.id}))
    return redirect('user_manage')


def user_manage_del_box(request):
    '''
    当选中用户时，删除用户
    :param request:
    :return:
    '''
    u_id = request.POST.get('u_id')
    u_id = u_id.split(',')


def user_start_use(request, user_id):
    '''
    启用用户
    :param request:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.user_status = '启用'
    user.save()
    return redirect(reverse(user_manage_edit, kwargs={'user_id': user_id}))
    # return redirect('start_and_stop')


def user_stop_use(request, user_id):
    '''
    禁用用户
    :param request:
    :param user_id:
    :return:
    '''
    user = UserInfo.objects.get(id=user_id)
    user.user_status = '禁用'
    user.save()
    # return redirect('user_manage')
    return redirect(reverse(user_manage_edit, kwargs={'user_id': user_id}))

    # return redirect('start_and_stop')


def start_and_stop(request):
    '''
    用户的启用与禁用
    :param request:
    :return:
    '''
    stop_users = UserInfo.objects.filter(user_status='禁用')  # 禁用
    start_users = UserInfo.objects.filter(user_status='启用')  # 启用
    return render(request, 'administrator/start_and_stop.html', {'stop_users': stop_users, 'start_users': start_users})


def start_and_stop_detail(request, num):
    '''
    启用与禁用详情
    :param request:
    :param num:
    :return:
    '''
    users = UserInfo.objects.filter(user_status=num)
    # return render(request, 'administrator/user_manage.html', {"users": users})
    return render(request, 'administrator/start_and_stop.html', {"users": users})


def administrator_password(request):
    '''
    密码管理
    :param request:
    :return:
    '''
    return render(request, 'administrator/administrator_password.html')
