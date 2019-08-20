from django.shortcuts import render, redirect, reverse, HttpResponse
from registration.models import *
from openpyxl import Workbook
import os
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import JsonResponse


def index(request):
    return render(request, 'pay_money_place/index.html')


def enter_index(request):
    """
    住院首页
    :param request:
    :return:
    """
    messages = PatientOne.objects.all()
    limit = 2
    paginator = Paginator(messages, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    return render(request, 'pay_money_place/enter_index.html', {'messages': result})


def enter_look(request, l_id):
    """
    查看住院详情
    :param request:
    :param l_id:
    :return:
    """
    each_data = PatientOne.objects.get(id=l_id)
    money = Money.objects.get(id=l_id)
    return render(request, 'pay_money_place/look.html', {'each_data': each_data, 'money': money})


def enter_edit(request, l_id):
    """
    编辑住院信息
    :param request:
    :param l_id:
    :return:
    """
    each_data = PatientOne.objects.get(id=l_id)
    enter_patient = Main_details.objects.get(id=l_id)
    all_depart = Department.objects.all()
    return render(request, 'pay_money_place/edit.html',
                  {'each_data': each_data, "enter_patient": enter_patient, "all_depart": all_depart})


def enter_add_many(request, l_id):
    """
    缴纳住院押金
    :param request:
    :param l_id:
    :return:
    """
    each_data = PatientOne.objects.get(id=l_id)
    return render(request, 'pay_money_place/add_many.html', {'each_data': each_data})


def outer_index(request):
    """
    出院结算
    :param request:
    :return:
    """
    money = Money.objects.all()
    limit = 1
    paginator = Paginator(money, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    # 余额
    pat_info = PatientOne.objects.all()
    for i in pat_info:
        smoney = Money.objects.get(fk_patientone=i.id)
        surplus = smoney.collect_money - int(i.patient_live_day) * 30
    return render(request, 'pay_money_place/outer_index.html', {'money': result, 'surplus': surplus})


def give_medicine_index(request):
    """
    在线发药首页
    :param request:
    :return:
    """
    all_data = Main_details.objects.all()
    limit = 2
    paginator = Paginator(all_data, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    return render(request, 'pay_money_place/give_medicine_index.html', {'all_data': result})


def pay_money_place_password(request):
    return render(request, 'pay_money_place/pay_money_place_password.html')


def enter_add(request):
    """
    住院结算
    :param request:
    :return:
    """
    if request.method == 'GET':
        return render(request, 'pay_money_place/enter_add.html')
    else:
        patient_card = request.POST.get('card')

        main_details = PatientOne.objects.get(id=patient_card)

        patient_money = request.POST.get('pmoney')
        print(patient_card)

        if patient_money != '':
            b = Medicine_one.objects.get(fk_patient_medicine_one_id=patient_card)
            medicine_name = b.medicine_one_name
            c = Medicine.objects.get(medicine_name=medicine_name)
            money_source = str(b.medicine_one_number) + b.medicine_one_name + '一共花费了' + str(
                b.medicine_one_number * b.medicine_one_outer_price)
            money_profit = b.medicine_one_number * b.medicine_one_outer_price - b.medicine_one_number * c.medicine_enter_price + int(
                patient_money)
            a = Money.objects.create(fk_patientone_id=patient_card, source=money_source, profit=money_profit,
                                     collect_money=int(patient_money), is_hospital=main_details.patient_is_live,
                                     collect_time=main_details.patient_one_outer_time)
        else:
            pass

    return render(request, 'pay_money_place/enter_add.html', {'main_details': main_details})


def outer_detail(request, o_id):
    """
    住院详情结算
    :param request:
    :param o_id:
    :return:
    """

    each_data = PatientOne.objects.get(id=o_id)
    return render(request, 'pay_money_place/outer_detail.html', {'each_data': each_data})


def give_detail(request, m_id):
    """
    发药详情
    :param request:
    :param m_id:
    :return:
    """
    each_data = Main_details.objects.get(id=m_id)
    return render(request, 'pay_money_place/give_detail.html', {'each_data': each_data})


def medicine_look(request, m_id):
    """
    查看药品数量
    :param request:
    :param m_id:
    :return:
    """
    each_data = Main_details.objects.get(id=m_id)
    med_name = each_data.drug_name
    if Medicine.objects.filter(medicine_name=med_name).exists():
        s = Medicine.objects.get(medicine_name=med_name)
        all_money = s.medicine_outer_price * each_data.drug_number
        return render(request, 'pay_money_place/medicine-look.html',
                      {'each_data': each_data, 's': s, 'all_money': all_money})
    else:
        return HttpResponse('此人的药没有了')


# 发药
def give_medicine(request, m_id):
    """
    发药功能实现
    :param request:
    :param m_id:
    :return:
    """
    each_data = Main_details.objects.get(id=m_id)
    patient_information_id = Main_details.objects.get(id=m_id)
    med_name = patient_information_id.drug_name
    med_number = patient_information_id.drug_number
    storage = Medicine.objects.get(medicine_name=med_name)
    finish_num = storage.medicine_number - med_number
    s = Medicine.objects.get(id=storage.id)
    print(s.medicine_outer_price)
    s.medicine_number = finish_num
    each_data.drug_status = 1
    s.save()
    each_data.save()
    all_money = s.medicine_outer_price * each_data.drug_number
    return redirect(reverse('medicine_look', kwargs={'m_id': m_id}))


# 在线发药查询
def inquire(request):
    """
    发药查询信息的功能
    :param request:
    :return:
    """
    gave = request.POST.get('pname')
    if Main_details.objects.filter(mediacal_redord_NO__contains=gave).exists():
        request.session['card'] = gave
        card = request.session.get('card')
        a = Main_details.objects.filter(mediacal_redord_NO__contains=gave)
        for i in a:
            return render(request, 'pay_money_place/give_detail.html', {'each_data': i})
    else:
        return HttpResponse('没有病历号')


# 更改病人信息
def updata_info(request, l_id):
    """
    更改住院病人信息的功能
    :param request:
    :param l_id:
    :return:
    """
    if request.method == "GET":
        each_data = PatientOne.objects.get(id=l_id)
        return render(request, 'pay_money_place/edit.html', {'each_data': each_data})
    else:
        main_doctor = request.POST.get('main_doctor')

        department = request.POST.getlist('department')[0]
        print(department)
        depart = Department.objects.get(depart_name=department)

        depart_id = depart.id

        each_data = PatientOne.objects.get(id=l_id)

        each_data.patient_main_doctor = main_doctor
        each_data.fk_patient_department_id = depart_id
        each_data.save()

        return redirect(reverse('look', kwargs={'l_id': l_id}))


# 缴纳押金
def cash_pledge(request, l_id):
    """
    缴纳押金的功能
    :param request:
    :param l_id:
    :return:
    """
    deposit = request.POST.get('pname')
    enter_patient = PatientOne.objects.all()
    each_data = PatientOne.objects.get(id=l_id)
    money = Money.objects.get(id=l_id)
    if money.collect_money >= 0:
        money.collect_money = int(deposit) + int(money.collect_money)

    else:
        return HttpResponse('已欠费')
    money.save()
    return render(request, 'pay_money_place/look.html', {'each_data': each_data, 'money': money})


# 结算
def close(request, o_id):
    """
    出院时将费用从未结算改为结算
    :param request:
    :param o_id:
    :return:
    """
    money = Money.objects.get(id=o_id)
    money.is_hospital = 1
    money.save()
    return redirect('outer_index')


# 导出住院办理表Excel表格
def export_excel(request):
    """
    导出住院办理Excel表格
    :param request:
    :return:
    """
    a = Workbook()
    sheet = a.active
    sheet.title = u'住院办理表'
    list_obj = PatientOne.objects.all()
    e = [[], ]
    for obj in list_obj:
        b = [obj.id, obj.patient_name, obj.patient_main_doctor, obj.patient_one_enter_time,
             obj.fk_patient_department_id, obj.patient_status]
        e.append(b)
    c = ['病历号', '姓名', '主治医生', '入院时间', '科室', '状态']
    e[0] = c
    for i in range(len(e)):
        for j in range(len(e[i])):
            sheet.cell(i + 1, j + 1, e[i][j])
    exits_file = os.path.exists('住院办理表.xlsx')
    if exits_file:
        os.remove(r'住院办理表.xlsx')
    a.save('d:\\住院办理表.xlsx')
    a = open('d:\\住院办理表.xlsx', 'rb')
    response = HttpResponse()
    response.content = a
    response['content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename=住院办理表.xls'
    return response


# 导出住院结算表
def export_excel_patient(request):
    """
    导出住院结算Excel表格
    :param request:
    :return:
    """
    a = Workbook()
    sheet = a.active
    # sheet = a.add_sheet(u'住院办理表')
    sheet.title = u'住院结算表'
    list_obj = Money.objects.all()
    e = [[], ]
    for obj in list_obj:
        b = [obj.fk_patientone.id, obj.fk_patientone.patient_name, obj.collect_money, obj.is_hospital]
        e.append(b)
    c = ['病历号', '姓名', '押金', '状态']
    e[0] = c
    print(e)
    for i in range(len(e)):
        for j in range(len(e[i])):
            sheet.cell(i + 1, j + 1, e[i][j])
    exits_file = os.path.exists('住院结算表.xlsx')
    if exits_file:
        os.remove(r'住院结算表.xlsx')
    a.save('d:\\住院结算表.xlsx')
    response = HttpResponse()
    a = open('d:\\住院结算表.xlsx', 'rb')
    response.content = a
    response['content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename=住院结算表.xls'
    return response


# 导住院结算详情表
def export_detail(request, o_id):
    """
    导出住院结算详情表
    :param request:
    :param o_id:
    :return:
    """
    a = Workbook()
    sheet = a.active
    # sheet = a.add_sheet(u'住院办理表')
    # sheet.title = u'结算详情表'
    list_obj = PatientOne.objects.get(id=o_id)
    b = [list_obj.id, list_obj.patient_name, list_obj.patient_medicine, list_obj.patient_live_day,
         list_obj.patient_one_enter_time]

    c = ['病历号', '姓名', '收费项目', '收费金额', '收费日期', []]
    c[1] = b
    for i in range(len(c)):
        for j in range(len(c[i])):
            sheet.cell(i + 1, j + 1, c[i][j])
    exits_file = os.path.exists('结算详情表.xlsx')
    if exits_file:
        os.remove(r'结算详情表.xlsx')
    a.save('d:\\结算详情表.xlsx')
    # a.close()
    a = open('d:\\住院详情表.xlsx', 'rb')
    response = HttpResponse()
    response.content = a
    response['content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename=住院详情表.xls'
    return response


# 住院查询
def hospitalization(request):
    """
    住院查询病历号
    :param request:
    :return:
    """
    card = request.POST.get('card_name')
    if card != None:
        if PatientOne.objects.filter(Q(id__contains=card) | Q(patient_name__icontains=card)).exists():
            request.session['a_card'] = card
            card = request.session.get('a_card')
            pat_info = PatientOne.objects.filter(Q(id__contains=card) | Q(patient_name__icontains=card))
            limit = 2
            paginator = Paginator(pat_info, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'pay_money_place/enter_index.html', {'messages': result})
        else:
            return HttpResponse('没有此病历号')
    else:
        return HttpResponse('请输入查询值')


# 住院结算查询
def query(request):
    """
    结算住院费用查询
    :param request:
    :return:
    """
    card = request.POST.get('pcard')
    if PatientOne.objects.filter(id=card).exists():
        request.session['card'] = card
        card = request.session.get('card')
        pat_info = PatientOne.objects.get(id=card)
        account = Money.objects.filter(fk_patientone=pat_info.id)
        limit = 1
        paginator = Paginator(account, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)

        return render(request, 'pay_money_place/outer_index.html', {'money': result})
    else:
        return HttpResponse('没有此病历号')


# 在线发药删除
def del_patient(request):
    nid = request.POST.getlist('check')[0]
    sid = nid.split(',')
    Main_details.objects.filter(id__in=sid).delete()
    return JsonResponse({'check': sid})


# 住院办理删除
def enter_del(request):
    nid = request.POST.getlist('check')[0]
    sid = nid.split(',')
    PatientOne.objects.filter(id__in=sid).delete()
    return JsonResponse({'check': sid})


# 删除出院信息
def outer_del(request):
    nid = request.POST.getlist('check')[0]
    sid = nid.split(',')
    Money.objects.filter(id__in=sid).delete()
    return JsonResponse({'check': sid})
